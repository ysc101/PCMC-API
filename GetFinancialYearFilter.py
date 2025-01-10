import requests
from openpyxl.reader.excel import load_workbook

from MyToken import bearer_token
import pandas as pd

url="http://192.168.0.162:8082/api/FinancialYearFilter/GetFinancialYearFilter"
headers={
    "Authorization": f"Bearer {bearer_token}"
}
params={
          "ID":"3"
}
response=requests.get(url, headers=headers,params=params)
if response.status_code==200:
    try:
        data=response.json()
        print("Response Data: ",data)
    except requests.exceptions.JSONDecodeError:
        print("Data Not in JSON")
    except requests.exceptions.HTTPError:
        print("URL is Invalid")
else:
    print("Response Code: ",response.status_code)
    print(response.text)

# Load the workbook and select the sheet
file_path = r"C:\Users\Yogesh.c\Desktop\LPRS_SNA_API_Document_Testing.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Find the next empty row in column A
next_row = sheet.max_row + 1

# Write data to the next empty row
sheet.cell(row=next_row, column="OutPut", value=data)

# Save the workbook
workbook.save(file_path)

print("Data appended to the next empty row.")