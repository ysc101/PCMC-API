from openpyxl import load_workbook

# Load the workbook and select the sheet
file_path = r"C:\Users\Yogesh.c\Desktop\LPRS_SNA_API_Document_Testing.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Find the next empty row in column A
next_row = sheet.max_row + 1

# Write data to the next empty row
sheet.cell(row=next_row, column=6, value="Appended Data")  # Appending in column A

# Save the workbook
workbook.save(file_path)

print("Data appended to the next empty row.")
